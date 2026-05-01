#!/usr/bin/env python3
"""
财经日报生成器 v2.0 - 演示版（无需API Key）
展示实时数据采集后的效果
"""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class FinanceDailyReportV2Demo:
    """财经日报生成器 v2.0 演示版 - 展示优化后的效果"""
    
    def __init__(self, company_name="公司名称", department="财务管理部 资金管理科", date=None):
        self.company_name = company_name
        self.department = department
        self.date = date or datetime.now().strftime("%Y年%m月%d日")
        
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "财经日报"
        
        self._init_styles()
    
    def _init_styles(self):
        """初始化样式"""
        self.thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        self.header_font = Font(name='微软雅黑', size=11, bold=True)
        self.title_font = Font(name='微软雅黑', size=14, bold=True, color='FF0000')
        self.section_font = Font(name='微软雅黑', size=11, bold=True)
        self.content_font = Font(name='微软雅黑', size=10)
        self.orange_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
        self.blue_fill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
        self.light_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    def get_demo_news_data(self) -> dict:
        """今日数据 - 2026年3月17日实时数据"""
        return {
            '资金动向': '''1、中国银行间债市表现平稳，央行开展逆回购操作维护流动性合理充裕。
2、USD/CNY中间价报6.8961，较前一交易日调升45个基点，市场汇率6.9004。
3、财政部今日公布国债发行计划，维持稳健货币政策基调。''',
            
            '宏观经济': '''1、中国2月官方制造业PMI为49.0%，环比下降0.3个百分点，制造业景气水平小幅回落。
2、国家统计局公布2月CPI数据，通胀水平维持低位运行。
3、2026年政府工作报告明确GDP增长目标5%左右，赤字率拟按4%左右安排。''',
            
            '国内新闻': '''1、国家能源局：推进农村能源革命，扩大农村充电设施覆盖，支持分布式光伏发展。
2、华为鸿蒙智行全系累计交付量达128万辆，发布新一代双光路激光雷达技术。
3、小米集团雷军提交建议案，涵盖人形机器人、智能驾驶安全准则等领域。''',
            
            '行业新闻': '''1、长城汽车2月海外销量同比增长35%，新能源汽车渗透率达38%。
2、国内动力电池2月装机量同比增长42%，磷酸铁锂电池占比超70%。
3、半导体行业协会：1-2月集成电路出口额同比增长18.5%。''',
            
            '俄乌相关': '''1、俄罗斯统计局：2025年实际工资增长4.4%，1月失业率降至2.2%历史低位。
2、美伊冲突持续升级：3月16日伊朗最高领袖哈梅内伊遭袭身亡，霍尔木兹海峡船舶流量下降93%。
3、布伦特原油价格突破103美元/桶，为2022年以来最高水平。''',
            
            '各国政策': '''1、美联储3月利率决议即将公布，市场预期维持利率不变，关注鲍威尔对经济前景的表态。
2、特朗普正式提名凯文·沃什出任下任美联储主席，需参议院确认。
3、美国财政部长：关税税率将恢复至此前水平，可能于本周实施15%全球关税。
4、欧元区1月失业率意外降至6.1%历史新低，经济复苏迹象显现。''',
            
            '国际新闻': '''1、美联储维持利率不变，鲍威尔表示通胀仍高于目标，将谨慎评估政策。
2、日本央行维持超宽松货币政策，日元汇率波动引发市场关注。
3、美伊冲突进入第三周，全球能源市场供应紧张，各国动用战略储备应对。''',
            
            '资本市场': '''1、美股：道指报48739点，标普500报6869点，纳指报22807点，受美伊冲突影响波动加大。
2、A股：沪指报3341点，深成指微涨，创业板指平盘，市场观望情绪浓厚。
3、港股：恒生指数报24231点，科技股表现分化，能源股受油价上涨提振。
4、欧洲：德国DAX涨1.74%，法国CAC40涨0.79%，英国富时100涨0.8%。''',
            
            '黄金原油': '''1、COMEX黄金期货报5061.70美元/盎司，地缘风险支撑避险需求，一度跌破5000美元后反弹。
2、布伦特原油报103.14美元/桶（+2.67%），霍尔木兹海峡封锁影响持续，日内最高触及103.95美元。
3、WTI原油报99.85美元/桶（+2.62%），美国考虑暂缓《琼斯法案》降低物流成本。
4、伊拉克准备削减约300万桶/日石油产量，全球能源供应面临严峻挑战。'''
        }
    
    def get_demo_policy_data(self) -> list:
        """演示政策数据"""
        return [
            {
                'dept': '财政部',
                'aspect': '国债发行',
                'policy': '拟于3月17日续发行320亿元50年期记账式附息国债，票面利率维持2.28%',
                'interpretation': '长期国债发行维持稳定利率，显示货币政策保持稳健，有利于企业锁定长期融资成本，降低财务费用。'
            },
            {
                'dept': '国家能源局',
                'aspect': '农村能源',
                'policy': '推进农村能源革命，提升电网供电保障，扩大农村充电设施覆盖，推进风电光伏开发利用',
                'interpretation': '利好新能源产业链，尤其是分布式光伏、充电桩、储能等相关企业，预计带动千亿级投资。'
            },
            {
                'dept': '工信部',
                'aspect': '智能驾驶',
                'policy': '加快明确L3/L4级智能驾驶安全准则，将L2级辅助驾驶"脱手脱眼"纳入交通违法处罚',
                'interpretation': '智能驾驶法规加速完善，有利于行业规范化发展，关注自动驾驶技术领先企业及产业链投资机会。'
            },
            {
                'dept': '央行',
                'aspect': '货币政策',
                'policy': '开展1000亿元逆回购操作，维护流动性合理充裕，保持货币政策稳健中性',
                'interpretation': '央行通过公开市场操作维护流动性，显示货币政策保持稳健，有利于金融市场稳定运行。'
            }
        ]
    
    def create_header(self):
        """创建表头"""
        ws = self.ws
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 25
        
        ws.merge_cells('A1:D1')
        ws['A1'] = self.company_name
        ws['A1'].font = Font(name='微软雅黑', size=12, bold=True)
        ws['A1'].alignment = self.center_align
        
        ws.merge_cells('A2:B2')
        ws['A2'] = '□ 绝密    □ 机密    □ 秘密    ☑ 一般'
        ws['A2'].font = Font(name='微软雅黑', size=9)
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells('C2:D2')
        ws['C2'] = f'{self.department}'
        ws['C2'].font = Font(name='微软雅黑', size=10)
        ws['C2'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws.merge_cells('A3:D3')
        ws['A3'] = self.date
        ws['A3'].font = Font(name='微软雅黑', size=10)
        ws['A3'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws.merge_cells('A4:D4')
        ws['A4'] = '今日财经新闻（日报）'
        ws['A4'].font = self.title_font
        ws['A4'].alignment = self.center_align
        
        return 5
    
    def create_focus_section(self, start_row, news_data):
        """创建聚焦热点部分"""
        ws = self.ws
        row = start_row
        
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = '一、聚焦热点'
        ws[f'A{row}'].font = self.section_font
        ws[f'A{row}'].fill = self.light_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 22
        row += 1
        
        headers = ['项目', '涉及方面', '新闻/政策', '']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.orange_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        ws.row_dimensions[row].height = 22
        row += 1
        
        categories = [
            ('资金动向', '资金动向'),
            ('宏观经济', '宏观经济'),
            ('国内新闻', '国内新闻'),
            ('行业新闻', '行业新闻'),
            ('俄乌相关', '俄乌相关'),
            ('各国政策', '各国政策'),
            ('国际新闻', '国际新闻'),
            ('资本市场', '资本市场'),
            ('黄金原油', '黄金原油'),
        ]
        
        for cat_name, cat_key in categories:
            content = news_data.get(cat_key, '')
            if not content:
                continue
            
            lines = content.count('\n') + 1
            row_height = max(60, lines * 15)
            
            ws.cell(row=row, column=1, value=cat_name)
            ws.cell(row=row, column=1).font = self.content_font
            ws.cell(row=row, column=1).alignment = self.center_align
            ws.cell(row=row, column=1).border = self.thin_border
            ws.cell(row=row, column=1).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            ws.cell(row=row, column=2, value='')
            ws.cell(row=row, column=2).font = self.content_font
            ws.cell(row=row, column=2).alignment = self.center_align
            ws.cell(row=row, column=2).border = self.thin_border
            
            ws.merge_cells(f'C{row}:D{row}')
            ws.cell(row=row, column=3, value=content)
            ws.cell(row=row, column=3).font = self.content_font
            ws.cell(row=row, column=3).alignment = self.left_align
            ws.cell(row=row, column=3).border = self.thin_border
            ws.cell(row=row, column=4).border = self.thin_border
            
            ws.row_dimensions[row].height = row_height
            row += 1
        
        return row
    
    def create_policy_section(self, start_row, policy_data):
        """创建国家政策部分"""
        ws = self.ws
        row = start_row
        
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = '二、国家政策'
        ws[f'A{row}'].font = self.section_font
        ws[f'A{row}'].fill = self.light_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 22
        row += 1
        
        headers = ['国家部委', '涉及方面', '新闻/政策', '政策解读']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.blue_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        ws.row_dimensions[row].height = 22
        row += 1
        
        for item in policy_data:
            ws.cell(row=row, column=1, value=item.get('dept', ''))
            ws.cell(row=row, column=1).font = self.content_font
            ws.cell(row=row, column=1).alignment = self.center_align
            ws.cell(row=row, column=1).border = self.thin_border
            ws.cell(row=row, column=1).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            ws.cell(row=row, column=2, value=item.get('aspect', ''))
            ws.cell(row=row, column=2).font = self.content_font
            ws.cell(row=row, column=2).alignment = self.center_align
            ws.cell(row=row, column=2).border = self.thin_border
            
            ws.cell(row=row, column=3, value=item.get('policy', ''))
            ws.cell(row=row, column=3).font = self.content_font
            ws.cell(row=row, column=3).alignment = self.left_align
            ws.cell(row=row, column=3).border = self.thin_border
            
            ws.cell(row=row, column=4, value=item.get('interpretation', ''))
            ws.cell(row=row, column=4).font = self.content_font
            ws.cell(row=row, column=4).alignment = self.left_align
            ws.cell(row=row, column=4).border = self.thin_border
            
            ws.row_dimensions[row].height = 60
            row += 1
        
        return row
    
    def generate(self, output_path=None):
        """生成演示日报"""
        print(f"\n{'='*60}")
        print(f"📰 财经日报 v2.0 演示版 - {self.date}")
        print(f"{'='*60}\n")
        
        print("✅ 使用演示数据（模拟实时采集效果）")
        print("   - 9大分类财经新闻")
        print("   - 4条国家政策及解读")
        print("   - 包含最新市场数据\n")
        
        news_data = self.get_demo_news_data()
        policy_data = self.get_demo_policy_data()
        
        row = self.create_header()
        row = self.create_focus_section(row, news_data)
        row = self.create_policy_section(row, policy_data)
        
        if output_path is None:
            output_dir = Path("/root/.openclaw/workspace/finance-reports")
            output_dir.mkdir(parents=True, exist_ok=True)
            date_str = datetime.now().strftime("%Y%m%d")
            output_path = output_dir / f"财经日报_{date_str}_v2_demo.xlsx"
        
        self.wb.save(output_path)
        
        print(f"{'='*60}")
        print(f"✅ 演示报告已生成: {output_path}")
        print(f"{'='*60}\n")
        
        print("💡 提示：这是演示版本，展示v2.0优化后的格式和数据结构")
        print("   正式版本需配置 Tavily API Key 实现实时数据采集")
        
        return output_path


if __name__ == "__main__":
    report = FinanceDailyReportV2Demo(
        company_name="公司名称",
        department="财务管理部 资金管理科"
    )
    report.generate()
