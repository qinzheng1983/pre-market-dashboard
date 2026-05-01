#!/usr/bin/env python3
"""
财经日报生成器
参照长城汽车日报样式，生成企业财务/资金管理部门的财经日报
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
from pathlib import Path

class FinanceDailyReport:
    def __init__(self, company_name="公司名称", department="财务管理部 资金管理科", date=None):
        self.company_name = company_name
        self.department = department
        self.date = date or datetime.now().strftime("%Y年%m月%d日")
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "财经日报"
        
        # 定义样式
        self._init_styles()
        
    def _init_styles(self):
        """初始化各种样式"""
        # 边框
        self.thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # 标题字体
        self.header_font = Font(name='微软雅黑', size=11, bold=True)
        self.title_font = Font(name='微软雅黑', size=14, bold=True, color='FF0000')
        self.section_font = Font(name='微软雅黑', size=11, bold=True)
        self.content_font = Font(name='微软雅黑', size=10)
        
        # 填充色
        self.orange_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
        self.blue_fill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
        self.light_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        
        # 对齐
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
    def create_header(self):
        """创建表头"""
        ws = self.ws
        
        # 设置列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 80
        ws.column_dimensions['D'].width = 25
        
        # 第1行：公司名称和密级
        ws.merge_cells('A1:D1')
        ws['A1'] = self.company_name
        ws['A1'].font = Font(name='微软雅黑', size=12, bold=True)
        ws['A1'].alignment = self.center_align
        
        # 第2行：密级选项
        ws.merge_cells('A2:B2')
        ws['A2'] = '□ 绝密    □ 机密    □ 秘密    ☑ 一般'
        ws['A2'].font = Font(name='微软雅黑', size=9)
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.merge_cells('C2:D2')
        ws['C2'] = f'{self.department}'
        ws['C2'].font = Font(name='微软雅黑', size=10)
        ws['C2'].alignment = Alignment(horizontal='right', vertical='center')
        
        # 第3行：日期
        ws.merge_cells('A3:D3')
        ws['A3'] = self.date
        ws['A3'].font = Font(name='微软雅黑', size=10)
        ws['A3'].alignment = Alignment(horizontal='right', vertical='center')
        
        # 第4行：主标题
        ws.merge_cells('A4:D4')
        ws['A4'] = '今日财经新闻（日报）'
        ws['A4'].font = self.title_font
        ws['A4'].alignment = self.center_align
        
        return 5  # 返回下一行号
        
    def create_focus_section(self, start_row, news_data):
        """创建聚焦热点部分"""
        ws = self.ws
        row = start_row
        
        # 一级标题：聚焦热点
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = '一、聚焦热点'
        ws[f'A{row}'].font = self.section_font
        ws[f'A{row}'].fill = self.light_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 22
        row += 1
        
        # 表头行
        headers = ['项目', '涉及方面', '新闻/政策', '']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.orange_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        ws.row_dimensions[row].height = 22
        row += 1
        
        # 数据映射
        categories = [
            ('资金动向', '资金动向'),
            ('宏观经济', '宏观经济'),
            ('国内新闻', '国内新闻'),
            ('行业新闻', '行业新闻'),
            ('俄乌相关', '俄乌相关'),
            ('各国政策', '各国政策'),
            ('国际新闻', '国际新闻'),
            ('资本市场', '资本市场'),
            ('黄金原油', '黄金原油'),
        ]
        
        for cat_name, cat_key in categories:
            content = news_data.get(cat_key, '')
            if not content:
                continue
                
            # 计算需要的行数（估算）
            lines = content.count('\n') + 1
            row_height = max(60, lines * 15)
            
            # 项目列（合并多行）
            ws.cell(row=row, column=1, value=cat_name)
            ws.cell(row=row, column=1).font = self.content_font
            ws.cell(row=row, column=1).alignment = self.center_align
            ws.cell(row=row, column=1).border = self.thin_border
            ws.cell(row=row, column=1).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # 涉及方面
            ws.cell(row=row, column=2, value='')
            ws.cell(row=row, column=2).font = self.content_font
            ws.cell(row=row, column=2).alignment = self.center_align
            ws.cell(row=row, column=2).border = self.thin_border
            
            # 新闻/政策内容
            ws.merge_cells(f'C{row}:D{row}')
            ws.cell(row=row, column=3, value=content)
            ws.cell(row=row, column=3).font = self.content_font
            ws.cell(row=row, column=3).alignment = self.left_align
            ws.cell(row=row, column=3).border = self.thin_border
            ws.cell(row=row, column=4).border = self.thin_border
            
            ws.row_dimensions[row].height = row_height
            row += 1
            
        return row
        
    def create_policy_section(self, start_row, policy_data):
        """创建国家政策部分"""
        ws = self.ws
        row = start_row
        
        # 一级标题：国家政策
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = '二、国家政策'
        ws[f'A{row}'].font = self.section_font
        ws[f'A{row}'].fill = self.light_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 22
        row += 1
        
        # 表头行
        headers = ['国家部委', '涉及方面', '新闻/政策', '政策解读']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.blue_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
        ws.row_dimensions[row].height = 22
        row += 1
        
        # 政策数据
        for item in policy_data:
            dept = item.get('dept', '')
            aspect = item.get('aspect', '')
            policy = item.get('policy', '')
            interpretation = item.get('interpretation', '')
            
            ws.cell(row=row, column=1, value=dept)
            ws.cell(row=row, column=1).font = self.content_font
            ws.cell(row=row, column=1).alignment = self.center_align
            ws.cell(row=row, column=1).border = self.thin_border
            ws.cell(row=row, column=1).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            ws.cell(row=row, column=2, value=aspect)
            ws.cell(row=row, column=2).font = self.content_font
            ws.cell(row=row, column=2).alignment = self.center_align
            ws.cell(row=row, column=2).border = self.thin_border
            
            ws.cell(row=row, column=3, value=policy)
            ws.cell(row=row, column=3).font = self.content_font
            ws.cell(row=row, column=3).alignment = self.left_align
            ws.cell(row=row, column=3).border = self.thin_border
            
            ws.cell(row=row, column=4, value=interpretation)
            ws.cell(row=row, column=4).font = self.content_font
            ws.cell(row=row, column=4).alignment = self.left_align
            ws.cell(row=row, column=4).border = self.thin_border
            
            ws.row_dimensions[row].height = 60
            row += 1
            
        return row
        
    def generate(self, news_data, policy_data, output_path=None):
        """生成完整的日报"""
        # 创建表头
        row = self.create_header()
        
        # 创建聚焦热点
        row = self.create_focus_section(row, news_data)
        
        # 创建国家政策
        row = self.create_policy_section(row, policy_data)
        
        # 保存文件
        if output_path is None:
            output_dir = Path("/root/.openclaw/workspace/finance-reports")
            output_dir.mkdir(parents=True, exist_ok=True)
            date_str = datetime.now().strftime("%Y%m%d")
            output_path = output_dir / f"财经日报_{date_str}.xlsx"
        
        self.wb.save(output_path)
        print(f"✅ 财经日报已生成: {output_path}")
        return output_path


def create_sample_report():
    """创建示例日报"""
    
    # 示例新闻数据（基于过去24小时的主要经济事件）
    news_data = {
        '资金动向': '''1、中国银行间债市表现强势，避险情绪主导之下债券向暖。利率债收益率纷纷下行，中短券表现较好。财政部拟3月11日第一次续发行320亿元50年期记账式附息国债。本次续发行国债票面利率与之前发行的同期国债相同，为2.28%。''',
        
        '宏观经济': '''1、十四届全国人大四次会议将于3月5日上午9时开幕，国务院总理李强将作《政府工作报告》，3月12日下午闭幕，会期8天，共安排3次全体会议。\n2、受春节假期等因素影响，中国2月官方制造业PMI为49.0%，环比下降0.3个百分点；非制造业PMI为49.5%，上升0.1个百分点；综合PMI产出指数为49.5%，下降0.3个百分点。''',
        
        '国内新闻': '''1、国家能源局召开会议强调，要深入推进农村能源革命，着力提升农村电网供电保障和综合承载能力，大力推进农村风电、光伏开发利用，扩大农村充电设施覆盖范围，持续做好北方地区清洁取暖，因地制宜推广农村可再生能源供暖，以能源发展带动村集体和村民增收。\n2、华为鸿蒙智行技术焕新发布会在深圳举办。发布会上，华为终端BG董事长余承东披露，截至3月3日，鸿蒙智行全系累计交付量已达128万辆。在技术层面，鸿蒙智行重点发布业界首创的新一代双光路图像级激光雷达技术，进一步强化全场景出行安全守护能力。与此同时，鸿蒙智行正式官宣与上汽联合打造的全新车型——尚界Z7与尚界Z7T。\n3、小米集团董事长雷军向大会提交五份建议案，涵盖人形机器人、智能汽车等不同领域。其中建议，扩大智能制造应用场景，提高人形机器人使用率；同时呼吁将L2级辅助驾驶"脱手脱眼"纳入交通违法处罚，尽快明确L3/L4级智驾安全准则。\n4、本田汽车宣布，计划将在中国生产的纯电动汽车（EV）进口至日本市场销售，预计最早于2026年春季正式推出。这是日本车企首次在日本本土进口并销售中国生产的纯电动车型。''',
        
        '行业新闻': '''1、长城汽车2月销售数据显示，海外市场销量持续增长，新能源汽车渗透率提升至35%。\n2、国内动力电池装机量2月同比增长42%，磷酸铁锂电池占比超过70%。''',
        
        '俄乌相关': '''1、俄罗斯统计局：2025年实际工资增长4.4%，1月份失业率为2.2%。\n2、俄罗斯经济部：1月国内生产总值（GDP）同比下降2.1%，前一个月同比增长1.9%。''',
        
        '各国政策': '''1、美国国防部长皮特·赫格塞斯表示，美伊冲突可能持续8周甚至更长时间，美方将掌控行动节奏与强度，北约拦截伊朗导弹不会触发集体防御条款。伊朗否认美媒有关伊方寻求与美国谈判的消息，最高领袖选举可能推迟到下周。\n2、美国财政部长贝森特表示，关税税率很快就会恢复到最高法院否决特朗普对等关税之前的水平。被问及美国何时正式采用15%的全球关税税率时，他表示"可能在本周的某个时候"。贝森特还表示，美国国际开发金融公司（DFC）将为在波斯湾地区运营的原油运输船及货船提供保险，相关安排已于3月3日宣布启动。\n3、美国总统特朗普正式提名凯文·沃什出任下一任美联储主席。若获得参议院确认，沃什将接替现任美联储主席鲍威尔职务，任期为四年。提名程序下一步为参议院银行业委员会听证会，但共和党参议员蒂利斯可能因鲍威尔接受刑事调查而阻止提名进入审议程序。\n4、欧元区1月失业率意外下滑至6.1%，创历史新低。此前分析师普遍预计该数据将维持不变。欧元区1月PPI同比下降2.1%，预期降2.7%，前值自降2.1%修正至降2%；环比升0.7%，预期升0.2%，前值降0.3%。''',
        
        '国际新闻': '''1、美联储3月利率决议即将公布，市场预期维持利率不变，但关注鲍威尔对经济前景的表态。\n2、日本央行行长表示将继续评估政策效果， yen汇率波动引发关注。''',
        
        '资本市场': '''美国三大股指全线收涨，道指涨0.49%报48739.41点，标普500指数涨0.78%报6869.5点，纳指涨1.29%报22807.48点。欧洲三大股指止跌反弹，德国DAX指数涨1.74%报24205.36点，法国CAC40指数涨0.79%报8167.73点，英国富时100指数涨0.8%报10567.65点。''',
        
        '黄金原油': '''国际贵金属期货普遍收涨，COMEX黄金期货涨0.54%报5151.60美元/盎司，COMEX白银期货涨0.35%报83.77美元/盎司。国际原油价格上涨，美油主力合约收涨2.08%，报76.11美元/桶；布油主力合约涨1.36%，报82.51美元/桶。随着储油空间即将耗尽，伊拉克开始关停其最大油田的石油生产。若霍尔木兹危机持续，该国还准备削减约300万桶/日的石油产量。'''
    }
    
    # 示例政策数据
    policy_data = [
        {
            'dept': '财政部',
            'aspect': '国债发行',
            'policy': '拟于3月11日续发行320亿元50年期记账式附息国债，票面利率维持2.28%',
            'interpretation': '长期国债发行维持稳定利率，显示货币政策保持稳健，有利于锁定长期融资成本。'
        },
        {
            'dept': '国家能源局',
            'aspect': '农村能源',
            'policy': '推进农村能源革命，提升电网供电保障，扩大农村充电设施覆盖',
            'interpretation': '利好新能源产业链，尤其是分布式光伏、充电桩等相关企业。'
        },
        {
            'dept': '工信部',
            'aspect': '智能驾驶',
            'policy': 'L2级辅助驾驶"脱手脱眼"拟纳入交通违法处罚，加快明确L3/L4级智驾安全准则',
            'interpretation': '智能驾驶法规加速完善，有利于行业规范化发展，关注自动驾驶技术领先企业。'
        }
    ]
    
    # 生成报告
    report = FinanceDailyReport(
        company_name="公司名称",
        department="财务管理部 资金管理科",
        date="2026年3月10日"
    )
    
    output_path = report.generate(news_data, policy_data)
    return output_path


if __name__ == "__main__":
    create_sample_report()
