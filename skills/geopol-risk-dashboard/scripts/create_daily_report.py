#!/usr/bin/env python3
"""
财经日报生成器 - 参照长城汽车日报样式
基于过去24小时国内外主要经济事件生成Excel日报
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# 创建输出目录
OUTPUT_DIR = "/root/.openclaw/workspace/daily-reports"
os.makedirs(OUTPUT_DIR, exist_ok=True)

def create_daily_report():
    """创建日报Excel文件"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "今日财经新闻"
    
    # 设置打印区域和页面设置
    ws.print_area = 'A1:F80'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = 9  # A4
    
    # 定义样式
    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    thick_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='medium'),
        bottom=Side(style='thin')
    )
    
    # 填充色
    header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')  # 深红色
    title_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')   # 淡黄色
    section_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')  # 橙色
    category_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # 浅绿色
    subcategory_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') # 浅黄色
    
    # 字体
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    title_font = Font(name='微软雅黑', size=14, bold=True, color='C00000')
    section_font = Font(name='微软雅黑', size=11, bold=True)
    normal_font = Font(name='微软雅黑', size=9)
    bold_font = Font(name='微软雅黑', size=9, bold=True)
    
    # 对齐方式
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # ========== 头部信息 ==========
    # 公司名称
    ws.merge_cells('A1:B1')
    ws['A1'] = '□ 绝密    □ 机密    □ 秘密    ☑ 一般'
    ws['A1'].font = Font(name='微软雅黑', size=9)
    ws['A1'].alignment = left_align
    
    ws.merge_cells('E1:F1')
    ws['E1'] = '财务管理部  资金管理科'
    ws['E1'].font = Font(name='微软雅黑', size=9, bold=True)
    ws['E1'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws.merge_cells('A2:B2')
    ws['A2'] = '×××××有限公司'
    ws['A2'].font = Font(name='微软雅黑', size=12, bold=True)
    ws['A2'].alignment = left_align
    
    ws.merge_cells('E2:F2')
    ws['E2'] = datetime.now().strftime('%Y年%m月%d日')
    ws['E2'].font = Font(name='微软雅黑', size=10)
    ws['E2'].alignment = Alignment(horizontal='right', vertical='center')
    
    # 大标题
    ws.merge_cells('A3:F3')
    ws['A3'] = '今日财经新闻（日报）'
    ws['A3'].font = title_font
    ws['A3'].alignment = center_align
    ws['A3'].fill = title_fill
    ws.row_dimensions[3].height = 30
    
    # ========== 一、聚焦热点 ==========
    ws.merge_cells('A4:F4')
    ws['A4'] = '一、聚焦热点'
    ws['A4'].font = section_font
    ws['A4'].alignment = left_align
    ws['A4'].fill = section_fill
    ws.row_dimensions[4].height = 22
    
    # 表头
    headers = ['项目', '涉及方面', '新闻政策']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col*2-1)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        if col == 1:
            ws.merge_cells(f'A5:B5')
        elif col == 2:
            ws.merge_cells(f'C5:D5')
        else:
            ws.merge_cells(f'E5:F5')
    
    # 数据内容
    row = 6
    
    # 资金动向
    ws.merge_cells(f'A{row}:A{row+1}')
    ws[f'A{row}'] = '资金动向'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = category_fill
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = '央行政策'
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].fill = subcategory_fill
    ws[f'B{row}'].border = thin_border
    
    ws.merge_cells(f'C{row}:F{row+1}')
    ws[f'C{row}'] = ('1、央行党委召开扩大会议强调，实施好适度宽松的货币政策，根据国内外经济金融形势和金融市场运行情况，择机降准降息，保持流动性充裕，使社会融资规模、货币供应量增长同经济增长、价格总水平预期目标相匹配。'
                    '2、央行行长潘功胜表示，2025年我国重点领域金融风险有序化解，融资平台债务风险化解取得重要阶段性成效，与2023年初相比融资平台数量和债务规模均下降超过70%。'
                    '3、预计2月新增贷款1.0-1.2万亿元，同比少增约2500-4500亿元，票据利率和同业存单利率之差大幅走扩至-57bps，银行体系面临"负债荒"和"资产荒"。')
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = top_left_align
    ws[f'C{row}'].border = thin_border
    
    row += 2
    
    # 宏观经济
    ws.merge_cells(f'A{row}:A{row+3}')
    ws[f'A{row}'] = '宏观经济'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = category_fill
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = '两会政策'
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].fill = subcategory_fill
    ws[f'B{row}'].border = thin_border
    
    ws.merge_cells(f'C{row}:F{row+3}')
    ws[f'C{row}'] = ('1、政府工作报告将2025年经济增长目标设定为"5%左右"，目标赤字率创新高，CPI涨幅目标2%左右。报告强调"稳住楼市股市"首次写入总体要求，政策层更加重视通过稳定资产价格发挥财富效应。'
                    '2、政府工作报告提出 "投资于人服务于民生"，加快推进部分品目消费税征收环节后移并下划地方，增加地方自主财力，开展中央部门零基预算改革试点。'
                    '3、政府工作报告提出"扎扎实实落实促进民营经济发展的政策措施"，切实依法保护民营企业和民营企业家合法权益，推动房地产止跌回稳，推进收购存量商品房。'
                    '4、国务院研究室表示，为达成通胀目标，今年要多措并举：加大逆周期调节力度，综合运用财政货币等宏观政策来改善供求关系；着力提振消费，释放需求侧潜力；综合整治"内卷式"竞争和价格战。')
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = top_left_align
    ws[f'C{row}'].border = thin_border
    
    row += 4
    
    # 国内新闻
    ws.merge_cells(f'A{row}:A{row+2}')
    ws[f'A{row}'] = '国内新闻'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = category_fill
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = '政策发布'
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].fill = subcategory_fill
    ws[f'B{row}'].border = thin_border
    
    ws.merge_cells(f'C{row}:F{row+2}')
    ws[f'C{row}'] = ('1、中共中央办公厅、国务院办公厅印发《提振消费专项行动方案》，提出加大生育养育保障力度，研究建立育儿补贴制度；加速推动自动驾驶、智能穿戴、超高清视频等新技术新产品开发与应用推广；严格落实带薪年休假制度，鼓励有条件的地方结合实际探索设置中小学春秋假。'
                    '2、国家发改委表示，中石油、中石化、中海油三大公司及其他原油加工企业要组织好成品油生产和调运，确保市场稳定供应，严格执行国家价格政策。'
                    '3、财政部预算报告显示，今年将发行3000亿元超长期特别国债用于支持消费品以旧换新，比上年增加1500亿元；将发行8000亿元超长期特别国债用于更大力度支持"两重"项目。')
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = top_left_align
    ws[f'C{row}'].border = thin_border
    
    row += 3
    
    # 行业新闻
    ws.merge_cells(f'A{row}:A{row+3}')
    ws[f'A{row}'] = '行业新闻'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = category_fill
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = '科技/制造业'
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].fill = subcategory_fill
    ws[f'B{row}'].border = thin_border
    
    ws.merge_cells(f'C{row}:F{row+3}')
    ws[f'C{row}'] = ('1、花旗集团将美国股市的评级从增持下调至中性，并将中国股市上调至增持。花旗认为，鉴于DeepSeek的人工智能技术突破、中国对科技行业的支持以及低估值，中国股市即使在最近的反弹之后也看起来很有吸引力。'
                    '2、近期全球铜价持续走高，市场担忧特朗普政府对铜征收关税的预期已引发美国进口抢购潮，花旗预计铜价将在3月份上涨至每吨10000美元。'
                    '3、国际能源署（IEA）发布《2025年电力报告》，预计2025-2027年将是全球电力需求增速最快的年份，年均增速保持在4%，新兴经济体将贡献85%的需求增长，其中中国占比超过50%。'
                    '4、国新办举行新闻发布会介绍"一揽子金融政策支持稳市场稳预期"有关情况，央行表示去年创设的两项支持资本市场工具首期额度分别是5000亿元、3000亿元，互换便利已开展2次操作、总金额1050亿元，超过500家上市公司公告使用贷款回购增持股票，贷款总金额约3000亿元。')
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = top_left_align
    ws[f'C{row}'].border = thin_border
    
    row += 4
    
    # 俄乌相关
    ws.merge_cells(f'A{row}:A{row+1}')
    ws[f'A{row}'] = '俄乌相关'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = category_fill
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = '俄乌局势'
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].fill = subcategory_fill
    ws[f'B{row}'].border = thin_border
    
    ws.merge_cells(f'C{row}:F{row+1}')
    ws[f'C{row}'] = ('1、俄罗斯总统普京与美国总统特朗普进行通话，双方通话重点是与伊朗相关的中东地区形势和乌克兰问题的谈判进程。'
                    '2、俄罗斯统计局数据显示，2025年实际工资增长4.4%，1月份失业率为2.2%；1月国内生产总值(GDP)同比下降2.1%，前一个月同比增长1.9%。')
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = top_left_align
    ws[f'C{row}'].border = thin_border
    
    row += 2
    
    # 各国政策
    ws.merge_cells(f'A{row}:A{row+3}')
    ws[f'A{row}'] = '各国政策'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = category_fill
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = '美国/欧盟'
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].fill = subcategory_fill
    ws[f'B{row}'].border = thin_border
    
    ws.merge_cells(f'C{row}:F{row+3}')
    ws[f'C{row}'] = ('1、美国总统特朗普表示美伊冲突可能持续8周甚至更长时间，美方将掌控行动节奏与强度，北约拦截伊朗导弹不会触发集体防御条款。特朗普还称"心中已有人选来接替哈梅内伊"。'
                    '2、美国财政部长贝森特表示，关税税率很快就会恢复到最高法院否决特朗普对等关税之前的水平，被问及美国何时正式采用15%的全球关税税率时，他表示"可能在本周的某个时候"。'
                    '3、欧盟理事会最终批准了一项增值税方案，旨在使欧盟增值税规则适应数字时代，该方案将提高欧盟的竞争力、帮助打击增值税欺诈，并减轻企业的行政负担。'
                    '4、欧洲央行行长拉加德表示，全球贸易和欧洲国防架构的突然转变将加大保持通胀稳定的难度，这些转变与气候变化一起构成了"双向冲击"。')
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = top_left_align
    ws[f'C{row}'].border = thin_border
    
    row += 4
    
    # 国际新闻
    ws.merge_cells(f'A{row}:A{row+2}')
    ws[f'A{row}'] = '国际新闻'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = category_fill
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = '国际动态'
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].fill = subcategory_fill
    ws[f'B{row}'].border = thin_border
    
    ws.merge_cells(f'C{row}:F{row+2}')
    ws[f'C{row}'] = ('1、欧盟委员会宣布将对美国的钢铝关税政策采取反制措施，特朗普对所有进口至美国的钢铁和铝征收25%关税的举措已正式生效，欧盟方案将从4月1日开始，并于4月13日全面到位。'
                    '2、加拿大央行决定将政策利率下调25个基点至2.75%，与预期值一致。'
                    '3、日本央行行长植田和男表示，随着进口成本驱动的通胀消退且薪资继续强劲增长，预计日本实际工资和消费者支出将改善，央行正在逐步缩减资产负债表规模。')
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = top_left_align
    ws[f'C{row}'].border = thin_border
    
    row += 3
    
    # 资本市场
    ws.merge_cells(f'A{row}:A{row+1}')
    ws[f'A{row}'] = '资本市场'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = category_fill
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = '全球股市'
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].fill = subcategory_fill
    ws[f'B{row}'].border = thin_border
    
    ws.merge_cells(f'C{row}:F{row+1}')
    ws[f'C{row}'] = ('美国三大股指全线重挫，道指跌2.08%收跌近900点；纳斯达克指数跌4%；标普500指数跌2.7%。大型科技股集体下跌，特斯拉暴跌15.43%创4年多来最大单日跌幅，英伟达跌5.07%，苹果跌4.85%，谷歌跌4.49%。中概股也未能幸免，纳斯达克中国金龙指数跌3.59%，阿里巴巴、理想汽车跌超5%。')
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = top_left_align
    ws[f'C{row}'].border = thin_border
    
    row += 2
    
    # 黄金原油
    ws.merge_cells(f'A{row}:A{row+2}')
    ws[f'A{row}'] = '黄金原油'
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].alignment = center_align
    ws[f'A{row}'].fill = category_fill
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = '大宗商品'
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = center_align
    ws[f'B{row}'].fill = subcategory_fill
    ws[f'B{row}'].border = thin_border
    
    ws.merge_cells(f'C{row}:F{row+2}')
    ws[f'C{row}'] = ('1、国际贵金属期货普遍收涨，COMEX黄金期货涨0.54%报5151.60美元/盎司，COMEX白银期货涨0.35%报83.77美元/盎司。'
                    '2、国际原油价格上涨，美油主力合约收涨2.08%报76.11美元/桶；布油主力合约涨1.36%报82.51美元/桶。随着储油空间即将耗尽，伊拉克开始关停其最大油田的石油生产，若霍尔木兹危机持续，该国还准备削减约300万桶/日的石油产量。'
                    '3、外交部发言人表示，能源安全对世界经济至关重要，各方都有责任确保能源供应稳定畅通，中方将采取必要措施保障自身能源安全。')
    ws[f'C{row}'].font = normal_font
    ws[f'C{row}'].alignment = top_left_align
    ws[f'C{row}'].border = thin_border
    
    row += 3
    
    # ========== 二、国家政策 ==========
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = '二、国家政策'
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].alignment = left_align
    ws[f'A{row}'].fill = section_fill
    ws.row_dimensions[row].height = 22
    
    row += 1
    # 表头
    policy_headers = ['国家部委', '涉及方面', '新闻政策', '政策解读']
    col_widths = [2, 2, 4, 4]  # 列跨度
    start_col = 1
    for i, header in enumerate(policy_headers):
        end_col = start_col + col_widths[i] - 1
        if col_widths[i] > 1:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        cell = ws.cell(row=row, column=start_col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        start_col = end_col + 1
    
    row += 1
    # 国家政策内容
    policy_data = [
        ('国务院关税税则委员会', '关税政策', 
         '自2025年3月10日起，对原产于美国的部分进口商品加征关税，对鸡肉、小麦、玉米、棉花加征15%关税；对高粱、大豆、猪肉、牛肉、水产品、水果、蔬菜、乳制品加征10%关税。',
         '针对美国对华加征的20%"芬太尼税"采取的反制措施，维护中国正当权益，平衡中美贸易关系。'),
        
        ('中国人民银行', '货币政策', 
         '实施好适度宽松的货币政策，择机降准降息，保持流动性充裕，使社会融资规模、货币供应量增长同经济增长、价格总水平预期目标相匹配。',
         '为经济持续向好创造适宜的货币金融环境，支持实体经济发展，稳定市场预期。'),
        
        ('国家发改委/财政部', '以旧换新', 
         '今年将发行3000亿元超长期特别国债用于支持消费品以旧换新，比上年增加1500亿元，资金已全部下达地方。',
         '扩内需、促消费的重要举措，推动消费升级和产业转型，提振市场信心。'),
        
        ('中共中央办公厅/国务院办公厅', '消费提振', 
         '印发《提振消费专项行动方案》，研究建立育儿补贴制度，严格落实带薪年休假制度，鼓励有条件的地方探索设置中小学春秋假。',
         '从民生角度出发提振消费，释放居民消费潜力，促进经济增长模式转型。'),
    ]
    
    for policy in policy_data:
        # 国家部委
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = policy[0]
        ws[f'A{row}'].font = bold_font
        ws[f'A{row}'].alignment = center_align
        ws[f'A{row}'].fill = category_fill
        ws[f'A{row}'].border = thin_border
        
        # 涉及方面
        ws[f'C{row}'] = policy[1]
        ws[f'C{row}'].font = normal_font
        ws[f'C{row}'].alignment = center_align
        ws[f'C{row}'].fill = subcategory_fill
        ws[f'C{row}'].border = thin_border
        
        # 新闻政策
        ws.merge_cells(f'D{row}:E{row}')
        ws[f'D{row}'] = policy[2]
        ws[f'D{row}'].font = normal_font
        ws[f'D{row}'].alignment = top_left_align
        ws[f'D{row}'].border = thin_border
        
        # 政策解读
        ws[f'F{row}'] = policy[3]
        ws[f'F{row}'].font = normal_font
        ws[f'F{row}'].alignment = top_left_align
        ws[f'F{row}'].border = thin_border
        
        ws.row_dimensions[row].height = 60
        row += 1
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 25
    
    # 保存文件
    filename = f"财经日报_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    
    return filepath

if __name__ == "__main__":
    filepath = create_daily_report()
    print(f"✅ 日报已生成: {filepath}")
